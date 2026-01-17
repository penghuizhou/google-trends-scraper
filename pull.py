from pytrends.request import TrendReq
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os
import time
import sys

def log_message(message):
    """Print timestamped log message"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"[{timestamp}] {message}")

def pull_google_trends():
    try:
        log_message("Starting Google Trends data pull...")
        
        log_message("Initializing connection to Google Trends...")
        pytrends = TrendReq(hl='en-US', tz=360)
        
        queries = ['Marble countertop', 'home remodel']
        log_message(f"Queries: {', '.join(queries)}")
        
        log_message("Waiting before request to avoid rate limiting...")
        time.sleep(5)
        
        log_message("Requesting data from Google Trends (timeframe: all-time, geo: US)...")
        pytrends.build_payload(queries, cat=0, timeframe='all', geo='US', gprop='')
        
        time.sleep(3)
        
        log_message("Fetching interest over time data...")
        df = pytrends.interest_over_time()
        
        if df.empty:
            log_message("ERROR: No data returned from Google Trends")
            return False
        
        log_message(f"Retrieved {len(df)} data points")
        
        if 'isPartial' in df.columns:
            df = df.drop('isPartial', axis=1)
        
        df = df.reset_index()
        df = df.rename(columns={'date': 'Month'})
        
        output_file = 'google_trends_data.xlsx'
        file_exists = os.path.exists(output_file)
        
        if file_exists:
            log_message("Existing file found. Updating with new data...")
            existing_df = pd.read_excel(output_file, sheet_name='Trends Data')
            df['Pull Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            combined_df = pd.concat([existing_df, df], ignore_index=True)
            original_count = len(combined_df)
            combined_df = combined_df.sort_values('Pull Date', ascending=False)
            combined_df = combined_df.drop_duplicates(subset=['Month'], keep='first')
            combined_df = combined_df.sort_values('Month')
            duplicates_removed = original_count - len(combined_df)
            if duplicates_removed > 0:
                log_message(f"Removed {duplicates_removed} duplicate entries")
            combined_df.to_excel(output_file, sheet_name='Trends Data', index=False)
        else:
            log_message("Creating new file...")
            df['Pull Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            df.to_excel(output_file, sheet_name='Trends Data', index=False)
        
        log_message("Formatting Excel file...")
        wb = load_workbook(output_file)
        ws = wb['Trends Data']
        
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        
        log_message("SUCCESS!")
        log_message(f"File saved to: {output_file}")
        log_message(f"Date range: {df['Month'].min()} to {df['Month'].max()}")
        log_message(f"Total records in file: {len(combined_df) if file_exists else len(df)}")
        
        return True
        
    except Exception as e:
        log_message(f"ERROR: {str(e)}")
        log_message(f"Error type: {type(e).__name__}")
        return False

        
if __name__ == "__main__":
    success = pull_google_trends()
    sys.exit(0 if success else 1)
